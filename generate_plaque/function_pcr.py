# generate_plaque/function_pcr.py
import pandas as pd
from itertools import zip_longest
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from io import BytesIO

def excel_coord_to_index(cell_ref: str) -> tuple[int, int]:
    col_letter, row = coordinate_from_string(cell_ref)
    col = column_index_from_string(col_letter)
    return row, col

def split_clean(s):
    return [t.strip() for t in str(s).split(";") if t.strip() != ""]

def dilution_sort_key(dil) -> tuple:
    """
    Ordre de tri pour une valeur de dilution :
      0 → SM (sérum pur, toujours en premier)
      1 → valeur numérique (croissant : 10, 100, 1000, …)
      2 → autres chaînes (ordre alphabétique)
    Retourne un tuple (tier, val) comparable par pandas / Python.
    """
    s = str(dil).strip() if dil is not None and str(dil).strip() != "" else ""
    if s.upper() == "SM":
        return (0, 0.0, "")
    try:
        return (1, float(s), "")
    except ValueError:
        return (2, 0.0, s)

def amorce_similarity_key(amorces_str: str, reference_sets: list[set]) -> tuple:
    """
    Retourne une clé de tri basée sur la similitude max avec les groupes déjà placés.
    Plus la valeur est grande, plus les amorces sont similaires au groupe dominant.
    On inverse pour trier décroissant (les plus similaires en premier).
    """
    my_set = set(split_clean(amorces_str))
    if not reference_sets:
        return (0,)
    max_common = max(len(my_set & ref) for ref in reference_sets)
    return (-max_common,)

def sort_by_primer_similarity(group: pd.DataFrame, group_dilutions: bool = False) -> pd.DataFrame:
    """
    Trie les échantillons par similitude d'amorces :
    1. Trouve le groupe d'amorces le plus fréquent → placé en premier
    2. Les suivants sont ceux qui partagent le plus d'amorces avec les déjà placés
    3. Les amorces totalement différentes vont à la fin

    Si group_dilutions=True, la dilution est triée en priorité dans chaque groupe
    d'amorces (avant le code labo), ce qui regroupe les échantillons de même dilution
    ensemble.
    """
    group = group.copy()

    # Convertir les amorces en sets pour comparaison
    group["_amorce_set"] = group["Amorces"].apply(lambda x: frozenset(split_clean(str(x))))

    # Compter les fréquences d'amorces exactes pour trouver le groupe dominant
    counts = group["_amorce_set"].value_counts()
    dominant = counts.index[0] if len(counts) > 0 else frozenset()

    # Score de similitude avec le groupe dominant
    group["_sim_score"] = group["_amorce_set"].apply(
        lambda s: -len(s & dominant)  # négatif pour tri décroissant
    )

    # Tri : similitude décroissante, puis amorces alphabétique
    # Si group_dilutions : Dilution avant Code labo (regrouper par dilution)
    # Sinon              : Code labo avant Dilution (comportement par défaut)
    if group_dilutions:
        sort_cols = ["_sim_score", "Amorces", "_dil_key", "Code labo", "Instance_num"]
    else:
        sort_cols = ["_sim_score", "Amorces", "Code labo", "Instance_num", "_dil_key"]

    group["_dil_key"] = group["Dilution"].apply(dilution_sort_key)
    sorted_group = group.sort_values(
        sort_cols,
        ascending=[True] * len(sort_cols)
    ).drop(columns=["_amorce_set", "_sim_score", "_dil_key"])

    return sorted_group

def sort_within_program_dil(group: pd.DataFrame) -> pd.DataFrame:
    counts = group["Amorces"].value_counts()
    group = group.copy()
    group["Amorce_count"] = group["Amorces"].map(counts)
    group["_dil_key"] = group["Dilution"].apply(dilution_sort_key)
    return (group.sort_values(
        ["Amorce_count", "Amorces", "_dil_key", "Code labo", "Instance_num"],
        ascending=[False, True, True, True, True]
    ).drop(columns=["Amorce_count", "_dil_key"]))

def sort_within_program(group: pd.DataFrame) -> pd.DataFrame:
    counts = group["Amorces"].value_counts()
    group = group.copy()
    group["Amorce_count"] = group["Amorces"].map(counts)
    group["_dil_key"] = group["Dilution"].apply(dilution_sort_key)
    return (group.sort_values(
        ["Amorce_count", "Amorces", "Code labo", "Instance_num", "_dil_key"],
        ascending=[False, True, True, True, True]
    ).drop(columns=["Amorce_count", "_dil_key"]))

def assign_plates_columns(group: pd.DataFrame, plate_size: int = 96) -> pd.DataFrame:
    """
    Place les échantillons dans les puits d'une plaque 96 (8 lignes × 12 colonnes).
    Règle : chaque nouvelle amorce (valeur distincte de la colonne "Amorces") commence
    obligatoirement à la rangée A de la prochaine colonne disponible.
    Les cases restantes de la colonne précédente sont laissées vides (pas de Well assigné).
    """
    ROWS_PER_COL = 8   # A→H
    COLS_PER_PLATE = plate_size // ROWS_PER_COL  # 12 pour une plaque 96

    rows_labels = list("ABCDEFGH")

    plate_list   = []
    row_list     = []
    col_list     = []
    well_list    = []

    current_plate = 1
    current_col   = 1   # 1-indexed dans la plaque courante
    current_row   = 0   # 0-indexed (0=A … 7=H)
    prev_amorce   = None

    def _next_col():
        """Avance d'une colonne, change de plaque si nécessaire."""
        nonlocal current_plate, current_col
        current_col += 1
        if current_col > COLS_PER_PLATE:
            current_plate += 1
            current_col = 1

    for _, sample in group.iterrows():
        amorce = sample["Amorces"]

        # Changement d'amorce → forcer le début d'une nouvelle colonne
        if amorce != prev_amorce and prev_amorce is not None:
            # Si la colonne précédente était exactement pleine (A→H),
            # l'amorce suivante doit sauter une colonne supplémentaire (colonne vide).
            prev_col_was_full = (current_row >= ROWS_PER_COL)
            _next_col()
            current_row = 0
            if prev_col_was_full:
                _next_col()

        prev_amorce = amorce

        # Débordement de colonne en cours de remplissage (>8 échantillons pour cette amorce)
        if current_row >= ROWS_PER_COL:
            current_col += 1
            current_row = 0
            if current_col > COLS_PER_PLATE:
                current_plate += 1
                current_col = 1

        plate_list.append(current_plate)
        row_list.append(rows_labels[current_row])
        col_list.append(current_col)
        well_list.append(rows_labels[current_row] + str(current_col).zfill(2))

        current_row += 1

    out = group.copy()
    out["PlateNbr"] = plate_list
    out["Row"]      = row_list
    out["Col"]      = col_list
    out["Well"]     = well_list
    return out

def read_excel_file(file_like) -> pd.DataFrame:
    """Lit l'Excel uploadé (objets FileStorage / file-like). Aucune écriture disque."""
    df = pd.read_excel(file_like, header=[0, 1, 2])
    df.columns = [
        "_".join([str(x).strip()
                  for x in tup
                  if not str(x).startswith("Unnamed") and str(x).strip() not in ("nan", "None", "")])
        for tup in df.columns.to_flat_index()
    ]
    # Trouver la colonne Echantillon (nom peut varier selon le multi-index aplati)
    echantillon_col_name = next(
        (c for c in df.columns if "echantillon" in c.lower() or c.strip().lower() == "echantillon"),
        None
    )
    base_cols = [
        "Code labo",
        "Instance #",
        "PRE-PCR-MON (0,1)_Amorces (Standard, Rep [replicateid])",
        "PRE-PCR-MON (0,1)_Programme PCR (Standard, Rep [replicateid])",
        "PRE-PCR-MON (0,1)_Enzyme utilise (Standard, Rep [replicateid])",
        "PRE-PCR-MON (0,1)_Facteur de dilution (Standard, Rep [replicateid])"
    ]
    if echantillon_col_name:
        base_cols.insert(0, echantillon_col_name)
    rename_map = {
        "PRE-PCR-MON (0,1)_Amorces (Standard, Rep [replicateid])": "Amorces",
        "PRE-PCR-MON (0,1)_Programme PCR (Standard, Rep [replicateid])": "ProgrammePCR",
        "PRE-PCR-MON (0,1)_Enzyme utilise (Standard, Rep [replicateid])": "Enzyme",
        "PRE-PCR-MON (0,1)_Facteur de dilution (Standard, Rep [replicateid])": "Dilution",
        "Instance #": "Instance",
    }
    if echantillon_col_name:
        rename_map[echantillon_col_name] = "Echantillon"
    df2 = df[base_cols].rename(columns=rename_map)
    df2["Amorces_liste"] = df2["Amorces"].apply(split_clean)
    df2["Programme_liste"] = df2["ProgrammePCR"].apply(split_clean)
    df2["Dilution_liste"] = df2["Dilution"].apply(split_clean)
    df2["pairs"] = df2.apply(
        lambda r: list(zip_longest(r["Amorces_liste"], r["Programme_liste"], fillvalue=pd.NA)),
        axis=1
    )
    macron = df2.explode("pairs", ignore_index=True)
    macron[["Amorces", "ProgrammePCR"]] = macron["pairs"].apply(pd.Series)
    macron = macron.explode("Dilution_liste", ignore_index=True)
    macron["Dilution"] = macron["Dilution_liste"].apply(pd.Series)
    macron["ProgrammePCR"] = macron.apply(
        lambda x: f"{x['ProgrammePCR']}_{x['Enzyme']}"
        if pd.notna(x["Enzyme"]) and x["Enzyme"] != ""
        else str(x["ProgrammePCR"]),
        axis=1
    )
    keep_cols = ["Code labo", "Instance", "Amorces", "ProgrammePCR", "Enzyme", "Dilution"]
    if "Echantillon" in macron.columns:
        keep_cols.append("Echantillon")
    result = macron[keep_cols].copy()
    # Colonne Demande = préfixe avant "-A" dans l'Echantillon (ex: "260203-00052-A0005" → "260203-00052")
    if "Echantillon" in result.columns:
        result["Demande"] = result["Echantillon"].astype(str).str.rsplit("-A", n=1).str[0]
    else:
        result["Demande"] = ""
    return result


def df_from_layout(layout: dict) -> pd.DataFrame:
    """
    Reconstruit un DataFrame pandas (même structure que read_excel_file) à partir
    d'un layout JSON. Utilisé pour recréer le pickle de session lors du chargement
    d'un plan sauvegardé, afin que /regroup puisse fonctionner.
    """
    rows = []
    for prog in layout.get("programmes", []):
        prog_name = prog.get("name", "")
        for plate in prog.get("plates", []):
            for well in plate.get("wells", {}).values():
                if not well or well.get("is_blank"):
                    continue
                rows.append({
                    "Code labo":  str(well.get("code_labo", "") or ""),
                    "Instance":   str(well.get("instance", "1") or "1"),
                    "Amorces":    str(well.get("amorces", "") or ""),
                    "ProgrammePCR": prog_name,
                    "Enzyme":     "",
                    "Dilution":   str(well.get("dilution", "") or ""),
                    "Echantillon": str(well.get("echantillon", "") or ""),
                    "Demande":    str(well.get("demande", "") or ""),
                })
    if not rows:
        return pd.DataFrame(columns=["Code labo", "Instance", "Amorces", "ProgrammePCR",
                                      "Enzyme", "Dilution", "Echantillon", "Demande"])
    df = pd.DataFrame(rows)
    # Dédupliquer — chaque échantillon (code_labo, instance, amorces, dilution) doit
    # apparaître une seule fois dans le df pour que /regroup ne le place pas en double.
    df = df.drop_duplicates(subset=["Code labo", "Instance", "Amorces", "Dilution"])
    return df


def make_content(x: pd.Series) -> str:
    code = str(x["Code labo"])
    amorce = str(x["Amorces"])
    instance = str(x["Instance"])
    dilution = str(x["Dilution"]).strip()
    parts = [code]
    if instance not in ("1", "", "nan"):
        parts.append(instance)
    parts.append(amorce)
    if dilution and dilution != "nan":
        parts.append(dilution)
    return "_".join(parts)


def prepare_plates_data(
    dataframe: pd.DataFrame,
    sort_by_similarity: bool = True,
    group_dilutions: bool = False,
    plate_size: int = 96,
) -> dict:
    """
    Prépare les données de placement dans les plaques et retourne un dict JSON-serialisable.

    Structure retournée :
    {
      "programmes": [
        {
          "name": "PCR1",
          "plates": [
            {
              "plate_nbr": 1,
              "wells": {
                "A01": {"id": "...", "content": "...", "code_labo": "...", "amorces": "...", ...},
                ...
              }
            }
          ]
        }
      ]
    }
    """
    df = dataframe.copy()
    # Filtrer les lignes sans programme PCR valide
    df = df[df["ProgrammePCR"].notna() & (df["ProgrammePCR"].astype(str).str.strip() != "nan") & (df["ProgrammePCR"].astype(str).str.strip() != "")]
    df["Dilution"] = df["Dilution"].fillna("").astype(str).replace("nan", "")
    df["Instance"] = df["Instance"].fillna("").astype(str).replace("nan", "")
    df["Instance_num"] = pd.to_numeric(df["Instance"], errors="coerce").fillna(1)
    if "Demande" not in df.columns:
        df["Demande"] = ""

    if sort_by_similarity:
        df_sorted = (
            df.groupby("ProgrammePCR", group_keys=False)
            .apply(sort_by_primer_similarity, group_dilutions=group_dilutions)
            .reset_index(drop=True)
        )
    elif group_dilutions:
        df_sorted = (
            df.groupby("ProgrammePCR", group_keys=False)
            .apply(sort_within_program_dil)
            .reset_index(drop=True)
        )
    else:
        df_sorted = (
            df.groupby("ProgrammePCR", group_keys=False)
            .apply(sort_within_program)
            .reset_index(drop=True)
        )

    plates = (
        df_sorted.groupby("ProgrammePCR", group_keys=False)
        .apply(assign_plates_columns, plate_size=plate_size)
        .reset_index(drop=True)
    )
    plates["Content"] = plates.apply(make_content, axis=1)

    result = {"programmes": []}
    for prog, prog_df in plates.groupby("ProgrammePCR", sort=True):
        prog_data = {"name": prog, "plates": []}
        for plate_nbr, plate_df in prog_df.groupby("PlateNbr", sort=True):
            wells = {}
            for _, row in plate_df.iterrows():
                well_key = row["Row"] + str(int(row["Col"])).zfill(2)
                wells[well_key] = {
                    "id": f"{prog}__{plate_nbr}__{well_key}",
                    "content": row["Content"],
                    "code_labo": str(row["Code labo"]),
                    "amorces": str(row["Amorces"]),
                    "programme": str(row["ProgrammePCR"]),
                    "dilution": str(row["Dilution"]),
                    "instance": str(row["Instance"]),
                    "demande": str(row.get("Demande", "")),
                }
            prog_data["plates"].append({"plate_nbr": int(plate_nbr), "wells": wells})
        result["programmes"].append(prog_data)

    return result


# Palette de 16 couleurs (fond de cellule) — miroir exact du CSS arrange.css
_AMORCE_PALETTE = [
    "dbeafe",  # 0  bleu
    "fce7f3",  # 1  rose vif
    "dcfce7",  # 2  vert
    "ffedd5",  # 3  orange
    "ede9fe",  # 4  violet
    "fef9c3",  # 5  jaune ocre
    "cffafe",  # 6  cyan
    "fee2e2",  # 7  rouge
    "f0fdf4",  # 8  vert sapin
    "fdf4ff",  # 9  mauve
    "fff7ed",  # 10 brun-orange
    "f0f9ff",  # 11 bleu ciel
    "fdf2f8",  # 12 fuchsia
    "ecfdf5",  # 13 menthe
    "f1f5f9",  # 14 gris ardoise
    "fff1f2",  # 15 cramoisi
]

_AMORCE_FILLS = [PatternFill(fill_type="solid", fgColor=c) for c in _AMORCE_PALETTE]


def build_amorce_color_map(layout: dict) -> dict[str, PatternFill]:
    """
    Reproduit la logique de buildAmorceColorMap() (utils.js) :
    trie alphabétiquement les valeurs `amorces` uniques sur l'ensemble
    du layout (tous programmes confondus) et leur assigne un index 0–15
    (cyclique). Retourne { amorces_str: PatternFill }.
    """
    amorces_set: set[str] = set()
    for prog_data in layout.get("programmes", []):
        for plate in prog_data.get("plates", []):
            for well in plate.get("wells", {}).values():
                if well and not well.get("is_blank") and well.get("amorces"):
                    amorces_set.add(well["amorces"])
    color_map: dict[str, PatternFill] = {}
    for idx, amorces in enumerate(sorted(amorces_set)):
        color_map[amorces] = _AMORCE_FILLS[idx % 16]
    return color_map


def generate_plaque_from_layout(
    layout: dict,
    template_file: str,
    position: str,
) -> BytesIO:
    """
    Génère le fichier Excel final à partir du layout JSON (potentiellement modifié par l'utilisateur).

    layout = {"programmes": [{"name": ..., "plates": [{"plate_nbr": ..., "wells": {"A01": {"content": ...}}}]}]}
    """
    row_start, col_start = excel_coord_to_index(position)
    wb = load_workbook(template_file)
    ws = wb["Feuil1"]

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    row_order = {ch: i for i, ch in enumerate("ABCDEFGH")}
    current_row = row_start

    amorce_color_map = build_amorce_color_map(layout)

    for prog_data in layout["programmes"]:
        prog = prog_data["name"]
        for plate_info in prog_data["plates"]:
            plate_nbr = plate_info["plate_nbr"]
            wells = plate_info["wells"]  # {"A01": {"content": ...}, ...}

            if not wells:
                continue

            # Déduire les lignes/colonnes utilisées
            rows_used = sorted({w[0] for w in wells.keys()}, key=lambda c: row_order.get(c, 99))
            max_col = max(int(w[1:]) for w in wells.keys())
            cols_used = list(range(1, max_col + 1))  # 1 → dernière colonne occupée (vides incluses)
            nb_cols = len(cols_used) + 1
            col = col_start

            # Titre
            title = f"{prog} — Plaque {plate_nbr}"
            ws.merge_cells(
                start_row=current_row, start_column=col,
                end_row=current_row, end_column=col + nb_cols - 1
            )
            ctitle = ws.cell(row=current_row, column=col, value=title)
            ctitle.font = Font(bold=True)
            ctitle.alignment = Alignment(horizontal="center")
            current_row += 1

            # En-têtes colonnes
            headers = [""] + cols_used
            for j, val in enumerate(headers):
                cell = ws.cell(row=current_row, column=col + j, value=val)
                cell.border = border
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            current_row += 1

            start_r = current_row
            for i, row_lbl in enumerate(rows_used):
                # En-tête ligne
                cell = ws.cell(row=start_r + i, column=col, value=row_lbl)
                cell.border = border
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                for j, col_num in enumerate(cols_used):
                    well_key = row_lbl + str(col_num).zfill(2)
                    well_data = wells.get(well_key, None)
                    val = well_data.get("content", None) if well_data else None
                    c = ws.cell(row=start_r + i, column=col + j + 1, value=val)
                    c.border = border
                    c.alignment = Alignment(horizontal="center")
                    # Puits blanc : italique + couleur grise
                    if well_data and well_data.get("is_blank"):
                        c.font = Font(italic=True, color="888888")
                    elif well_data and well_data.get("amorces"):
                        fill = amorce_color_map.get(well_data["amorces"])
                        if fill:
                            c.fill = fill

            end_r = start_r + len(rows_used) - 1
            end_c = col + nb_cols - 1

            # Largeurs colonnes
            for c_idx in range(col, end_c + 1):
                max_len = 0
                for row_cells in ws.iter_rows(min_row=start_r, max_row=end_r, min_col=c_idx, max_col=c_idx):
                    for cell in row_cells:
                        if cell.value is not None:
                            max_len = max(max_len, len(str(cell.value)))
                col_letter = get_column_letter(c_idx)
                target = 6 if c_idx == col else max(max_len + 2, 12)
                current_w = ws.column_dimensions[col_letter].width or 0
                ws.column_dimensions[col_letter].width = max(current_w, target)

            current_row = end_r + 1 + 2

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_plaque_in_template(dataframe: pd.DataFrame, template_file: str, position: str, choice: bool) -> BytesIO:
    """
    Compatibilité ascendante avec l'ancienne route /upload.
    """
    layout = prepare_plates_data(dataframe, sort_by_similarity=False, group_dilutions=choice)
    return generate_plaque_from_layout(layout, template_file, position)


def prepare_plates_data_grouped(
    df: pd.DataFrame,
    groups: dict,
    sort_by_similarity: bool = True,
    group_dilutions: bool = False,
    plate_size: int = 96,
) -> dict:
    """
    Génère un layout groupé : chaque groupe de demandes produit ses propres plaques.

    groups = {"Groupe A": ["260203-00052", "260203-00034"], "Groupe B": [...]}

    Les demandes non présentes dans groups sont ignorées.
    Les noms de programme sont préfixés par le nom du groupe : "Groupe A — 55 35_TypeIT"
    """
    all_programmes = []

    for group_name, demandes in groups.items():
        if not demandes:
            continue
        df_group = df[df["Demande"].isin(demandes)].copy()
        if df_group.empty:
            continue
        partial = prepare_plates_data(df_group, sort_by_similarity, group_dilutions, plate_size)
        for prog_data in partial["programmes"]:
            prog_data["name"] = f"{group_name} \u2014 {prog_data['name']}"
            all_programmes.append(prog_data)

    return {"programmes": all_programmes}
